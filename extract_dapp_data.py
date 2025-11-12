from openpyxl import load_workbook
import json

wb = load_workbook('/Users/erinmagennis/Documents/personal/Test projects/ef_dapps_data.xlsx')

# Categories to include
categories = ['DeFi', 'Social', 'Privacy', 'Collectibles', 'Gaming', 'DAO', 'Productivity', 'Bridge']

all_data = {}

for category in categories:
    if category in wb.sheetnames:
        ws = wb[category]

        # Get headers
        headers = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(1, col)
            headers.append(cell.value)

        # Find column indices
        name_idx = headers.index('Name') + 1 if 'Name' in headers else 1
        url_idx = headers.index('URL') + 1 if 'URL' in headers else 2
        logo_idx = headers.index('Logo Image') + 1 if 'Logo Image' in headers else 4
        desc_idx = headers.index('Description') + 1 if 'Description' in headers else 3

        # Extract data
        category_data = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, name_idx).value
            url = ws.cell(row, url_idx).value
            logo = ws.cell(row, logo_idx).value
            desc = ws.cell(row, desc_idx).value

            # Only include if we have a name
            if name:
                category_data.append({
                    'name': name,
                    'url': url,
                    'logo': logo,
                    'description': desc
                })

        all_data[category] = category_data
        print(f'{category}: {len(category_data)} apps')

# Save to JSON
with open('/Users/erinmagennis/Documents/personal/Test projects/dapp_data.json', 'w') as f:
    json.dump(all_data, f, indent=2, default=str)

print(f'\nTotal apps across all categories: {sum(len(apps) for apps in all_data.values())}')
print('Data saved to dapp_data.json')
